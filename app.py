from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from datetime import datetime
import pandas as pd
import json
import requests
import os
from werkzeug.security import generate_password_hash, check_password_hash
import secrets

app = Flask(__name__)
CORS(app)

# Configurações
app.config['SECRET_KEY'] = secrets.token_hex(16)
DATA_FILE = 'dados_producao.json'
BORDADORES_FILE = 'bordadores.json'
USERS_FILE = 'usuarios.json'

@app.route('/')
def index():
    return render_template('index.html')

# Rotas de Botões 

@app.route('/gerenciamento')
def gerenciamento():
    return render_template('gerenciamento.html')  # ou 'gerenciamento.html' se for outro arquivo

@app.route('/pesquisa')
def pesquisa():
    return render_template('pesquisa.html')


# Funções auxiliares
def carregar_dados():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def salvar_dados(dados):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

def carregar_bordadores():
    if os.path.exists(BORDADORES_FILE):
        with open(BORDADORES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return ['João Silva', 'Maria Santos', 'Pedro Oliveira', 'Ana Costa']

def salvar_bordadores(bordadores):
    with open(BORDADORES_FILE, 'w', encoding='utf-8') as f:
        json.dump(bordadores, f, ensure_ascii=False, indent=2)

def carregar_usuarios():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    # Usuários padrão
    usuarios = {
        'gerente': {
            'senha': generate_password_hash('admin123'),
            'tipo': 'gerente',
            'nome': 'Gerente'
        },
        'colaborador': {
            'senha': generate_password_hash('colab123'),
            'tipo': 'colaborador',
            'nome': 'Colaborador'
        }
    }
    salvar_usuarios(usuarios)
    return usuarios

def salvar_usuarios(usuarios):
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(usuarios, f, ensure_ascii=False, indent=2)

# Rotas de Autenticação
@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    usuarios = carregar_usuarios()
    
    username = data.get('username')
    senha = data.get('senha')
    
    if username in usuarios and check_password_hash(usuarios[username]['senha'], senha):
        return jsonify({
            'success': True,
            'tipo': usuarios[username]['tipo'],
            'nome': usuarios[username]['nome'],
            'username': username
        })
    
    return jsonify({'success': False, 'message': 'Credenciais inválidas'}), 401


# # Rotas de Bordadores
@app.route('/api/bordadores', methods=['GET'])
def get_bordadores():
    return jsonify(carregar_bordadores())

@app.route('/api/bordadores', methods=['POST'])
def add_bordador():
    data = request.json
    bordadores = carregar_bordadores()
    novo_bordador = data.get('nome')
    
    if novo_bordador and novo_bordador not in bordadores:
        bordadores.append(novo_bordador)
        salvar_bordadores(bordadores)
        return jsonify({'success': True, 'bordadores': bordadores})
    
    return jsonify({'success': False, 'message': 'Bordador já existe ou nome inválido'}), 400

@app.route('/api/bordadores/<nome>', methods=['PUT'])
def update_bordador(nome):
    data = request.json
    bordadores = carregar_bordadores()
    novo_nome = data.get('nome')
    
    if not novo_nome or novo_nome.strip() == '':
        return jsonify({'success': False, 'message': 'Nome inválido'}), 400
    
    # Verificar se o nome antigo existe
    if nome not in bordadores:
        return jsonify({'success': False, 'message': 'Bordador não encontrado'}), 404
    
    # Verificar se o novo nome já existe (e não é o mesmo)
    if novo_nome in bordadores and novo_nome != nome:
        return jsonify({'success': False, 'message': 'Nome já existe'}), 400
    
    # Atualizar o nome
    index = bordadores.index(nome)
    bordadores[index] = novo_nome
    salvar_bordadores(bordadores)
    
    # Atualizar também nos registros de produção
    dados = carregar_dados()
    for registro in dados:
        if registro.get('Bordador') == nome:
            registro['Bordador'] = novo_nome
    salvar_dados(dados)
    
    # Atualizar também nos usuários (se houver usuário com esse nome)
    usuarios = carregar_usuarios()
    for username, user_data in usuarios.items():
        if user_data.get('nome') == nome:
            user_data['nome'] = novo_nome
    salvar_usuarios(usuarios)
    
    return jsonify({'success': True, 'bordadores': bordadores})

@app.route('/api/bordadores/<nome>', methods=['DELETE'])
def delete_bordador(nome):
    bordadores = carregar_bordadores()
    
    if nome not in bordadores:
        return jsonify({'success': False, 'message': 'Bordador não encontrado'}), 404
    
    # Verificar se há registros de produção para este bordador
    dados = carregar_dados()
    tem_registros = any(d.get('Bordador') == nome for d in dados)
    
    if tem_registros:
        return jsonify({
            'success': False, 
            'message': 'Não é possível excluir. Este bordador possui registros de produção.'
        }), 400
    
    # Remover o bordador
    bordadores.remove(nome)
    salvar_bordadores(bordadores)
    
    return jsonify({'success': True, 'bordadores': bordadores})

# Rotas de Produção
@app.route('/api/producao', methods=['GET'])
def get_producao():
    dados = carregar_dados()
    bordador = request.args.get('bordador')
    
    if bordador:
        dados = [d for d in dados if d.get('Bordador') == bordador]
    
    return jsonify(dados)

@app.route('/api/producao', methods=['POST'])
def add_producao():
    data = request.json
    dados = carregar_dados()
    
    # Adicionar ID único e timestamp
    novo_registro = {
        'id': len(dados) + 1,
        'timestamp': datetime.now().isoformat(),
        **data
    }
    
    dados.append(novo_registro)
    salvar_dados(dados)
    
    return jsonify({'success': True, 'registro': novo_registro})

@app.route('/api/producao/<int:id>', methods=['PUT'])
def update_producao(id):
    data = request.json
    dados = carregar_dados()
    
    for i, registro in enumerate(dados):
        if registro.get('id') == id:
            dados[i] = {**registro, **data, 'id': id}
            salvar_dados(dados)
            return jsonify({'success': True, 'registro': dados[i]})
    
    return jsonify({'success': False, 'message': 'Registro não encontrado'}), 404

@app.route('/api/producao/<int:id>', methods=['DELETE'])
def delete_producao(id):
    dados = carregar_dados()
    dados = [d for d in dados if d.get('id') != id]
    salvar_dados(dados)
    return jsonify({'success': True})

# Rota de Exportação
@app.route('/api/exportar', methods=['POST'])
def exportar_excel():
    import zipfile
    from io import BytesIO
    
    dados = carregar_dados()
    
    if not dados:
        return jsonify({'success': False, 'message': 'Sem dados para exportar'}), 400
    
    df_completo = pd.DataFrame(dados)
    
    # Remover colunas técnicas
    colunas_remover = ['id', 'timestamp']
    df_completo = df_completo.drop(columns=[col for col in colunas_remover if col in df_completo.columns])
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Criar um buffer de memória para o ZIP
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # 1. Planilha completa
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df_completo.to_excel(writer, index=False, sheet_name='Produção')
            worksheet = writer.sheets['Produção']
            
            # Ajustar larguras
            for idx, col in enumerate(df_completo.columns, 1):
                max_length = max(df_completo[col].astype(str).apply(len).max(), len(col)) + 2
                col_letter = chr(64 + idx) if idx <= 26 else f"A{chr(64 + idx - 26)}"
                worksheet.column_dimensions[col_letter].width = min(max_length, 30)
            
            # Formatar cabeçalho
            from openpyxl.styles import Font, PatternFill, Alignment
            
            for cell in worksheet[1]:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        excel_buffer.seek(0)
        zip_file.writestr(f'producao_completa_{timestamp}.xlsx', excel_buffer.read())
        
        # 2. Planilha por bordador
        bordadores = df_completo['Bordador'].unique()
        
        for bordador in bordadores:
            df_bordador = df_completo[df_completo['Bordador'] == bordador]
            
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df_bordador.to_excel(writer, index=False, sheet_name='Produção')
                worksheet = writer.sheets['Produção']
                
                # Ajustar larguras
                for idx, col in enumerate(df_bordador.columns, 1):
                    max_length = max(df_bordador[col].astype(str).apply(len).max(), len(col)) + 2
                    col_letter = chr(64 + idx) if idx <= 26 else f"A{chr(64 + idx - 26)}"
                    worksheet.column_dimensions[col_letter].width = min(max_length, 30)
                
                # Formatar cabeçalho
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            excel_buffer.seek(0)
            filename = f'{bordador.replace(" ", "_")}_{timestamp}.xlsx'
            zip_file.writestr(filename, excel_buffer.read())
    
    zip_buffer.seek(0)
    
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'producao_completa_{timestamp}.zip'
    )

# Rota de Estatísticas
@app.route('/api/estatisticas', methods=['GET'])
def get_estatisticas():
    dados = carregar_dados()
    bordador = request.args.get('bordador')
    
    if bordador:
        dados = [d for d in dados if d.get('Bordador') == bordador]
    
    total_registros = len(dados)
    
    # Calcular totais
    total_pecas = 0
    total_pontos = 0
    
    for d in dados:
        try:
            qtd = int(''.join(filter(str.isdigit, str(d.get('QTD', '0')))))
            total_pecas += qtd
        except:
            pass
        
        try:
            pontos = int(''.join(filter(str.isdigit, str(d.get('PONTOS', '0')))))
            total_pontos += pontos
        except:
            pass
    
    return jsonify({
        'total_registros': total_registros,
        'total_pecas': total_pecas,
        'total_pontos': total_pontos
    })


# Adicione esta rota no seu app.py, após a rota /api/estatisticas

@app.route('/api/buscar-pedido/<pedido_id>', methods=['GET'])
def buscar_pedido(pedido_id):
    """Busca todos os registros de um pedido específico e agrupa por bordador e posição"""
    dados = carregar_dados()
    
    # Filtrar registros pelo ID do pedido
    registros = [d for d in dados if d.get('ID') == pedido_id]
    
    if not registros:
        return jsonify({
            'success': False,
            'message': 'Nenhum registro encontrado para este ID'
        }), 404
    
    # Organizar informações por posição
    resultado = {
        'id_pedido': pedido_id,
        'total_registros': len(registros),
        'bordadores': {},
        'posicoes': {
            'FRENTE': [],
            'LATERAL': [],
            'TRASEIRA': []
        },
        'tipos': {
            'BONE': [],
            'CUMBUCA': [],
            'VISEIRA': []
        },
        'processos': {
            'BORDADO': [],
            'AP_PINT': [],
            'AP_GRAV': []
        },
        'resumo': {
            'total_pecas': 0,
            'total_pontos': 0
        }
    }
    
    # Processar cada registro
    for registro in registros:
        bordador = registro.get('Bordador', 'Desconhecido')
        
        # Inicializar bordador se não existe
        if bordador not in resultado['bordadores']:
            resultado['bordadores'][bordador] = {
                'registros': 0,
                'pecas': 0,
                'pontos': 0,
                'posicoes': [],
                'data': registro.get('Data', '')
            }
        
        # Contar registros e somar totais
        resultado['bordadores'][bordador]['registros'] += 1
        
        try:
            qtd = int(''.join(filter(str.isdigit, str(registro.get('QTD', '0')))))
            resultado['bordadores'][bordador]['pecas'] += qtd
            resultado['resumo']['total_pecas'] += qtd
        except:
            pass
        
        try:
            pontos = int(''.join(filter(str.isdigit, str(registro.get('PONTOS', '0')))))
            resultado['bordadores'][bordador]['pontos'] += pontos
            resultado['resumo']['total_pontos'] += pontos
        except:
            pass
        
        # Verificar posições
        if registro.get('FRENTE') == 'X':
            resultado['posicoes']['FRENTE'].append({
                'bordador': bordador,
                'qtd': registro.get('QTD', '0'),
                'pontos': registro.get('PONTOS', '0'),
                'data': registro.get('Data', ''),
                'bordado': registro.get('BORDADO', '') == 'X'
            })
            if 'Frente' not in resultado['bordadores'][bordador]['posicoes']:
                resultado['bordadores'][bordador]['posicoes'].append('Frente')
        
        if registro.get('LATERAL') == 'X':
            resultado['posicoes']['LATERAL'].append({
                'bordador': bordador,
                'qtd': registro.get('QTD', '0'),
                'pontos': registro.get('PONTOS', '0'),
                'data': registro.get('Data', ''),
                'bordado': registro.get('BORDADO', '') == 'X'
            })
            if 'Lateral' not in resultado['bordadores'][bordador]['posicoes']:
                resultado['bordadores'][bordador]['posicoes'].append('Lateral')
        
        if registro.get('TRASEIRA') == 'X':
            resultado['posicoes']['TRASEIRA'].append({
                'bordador': bordador,
                'qtd': registro.get('QTD', '0'),
                'pontos': registro.get('PONTOS', '0'),
                'data': registro.get('Data', ''),
                'bordado': registro.get('BORDADO', '') == 'X'
            })
            if 'Traseira' not in resultado['bordadores'][bordador]['posicoes']:
                resultado['bordadores'][bordador]['posicoes'].append('Traseira')
        
        # Verificar tipos
        if registro.get('BONE') == 'X':
            if bordador not in [t['bordador'] for t in resultado['tipos']['BONE']]:
                resultado['tipos']['BONE'].append({'bordador': bordador})
        
        if registro.get('CUMBUCA') == 'X':
            if bordador not in [t['bordador'] for t in resultado['tipos']['CUMBUCA']]:
                resultado['tipos']['CUMBUCA'].append({'bordador': bordador})
        
        if registro.get('VISEIRA') == 'X':
            if bordador not in [t['bordador'] for t in resultado['tipos']['VISEIRA']]:
                resultado['tipos']['VISEIRA'].append({'bordador': bordador})
        
        # Verificar processos
        if registro.get('BORDADO') == 'X':
            if bordador not in [p['bordador'] for p in resultado['processos']['BORDADO']]:
                resultado['processos']['BORDADO'].append({'bordador': bordador})
        
        if registro.get('AP_PINT') == 'X':
            if bordador not in [p['bordador'] for p in resultado['processos']['AP_PINT']]:
                resultado['processos']['AP_PINT'].append({'bordador': bordador})
        
        if registro.get('AP_GRAV') == 'X':
            if bordador not in [p['bordador'] for p in resultado['processos']['AP_GRAV']]:
                resultado['processos']['AP_GRAV'].append({'bordador': bordador})
    
    return jsonify({
        'success': True,
        'resultado': resultado
    })


# Rota para Gerente adicionar usuários
@app.route('/api/usuarios', methods=['POST'])
def add_usuario():
    data = request.json
    usuarios = carregar_usuarios()

    username = data.get('username')
    senha = data.get('senha')
    nome = data.get('nome')
    tipo = data.get('tipo', 'colaborador')

    if not username or not senha or not nome:
        return jsonify({'success': False, 'message': 'Dados incompletos'}), 400

    if username in usuarios:
        return jsonify({'success': False, 'message': 'Usuário já existe'}), 400

    usuarios[username] = {
        'senha': generate_password_hash(senha),
        'tipo': tipo,
        'nome': nome
    }

    salvar_usuarios(usuarios)
    
    # Se for colaborador, adicionar automaticamente na lista de bordadores
    if tipo == 'colaborador':
        bordadores = carregar_bordadores()
        if nome not in bordadores:
            bordadores.append(nome)
            salvar_bordadores(bordadores)

    return jsonify({'success': True, 'usuario': {
        'username': username,
        'nome': nome,
        'tipo': tipo
    }})


# Rota para listar todos os usuários
@app.route('/api/usuarios', methods=['GET'])
def get_usuarios():
    usuarios = carregar_usuarios()
    # Remover senhas antes de enviar
    usuarios_safe = {}
    for username, data in usuarios.items():
        usuarios_safe[username] = {
            'nome': data.get('nome'),
            'tipo': data.get('tipo')
        }
    return jsonify(usuarios_safe)

# Rota para atualizar usuário (já existe POST, agora adicionar PUT)
@app.route('/api/usuarios/<username>', methods=['PUT'])
def update_usuario(username):
    data = request.json
    usuarios = carregar_usuarios()
    
    if username not in usuarios:
        return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404
    
    # Atualizar dados
    if 'nome' in data and data['nome']:
        old_name = usuarios[username]['nome']
        usuarios[username]['nome'] = data['nome']
        
        # Atualizar nome nos bordadores se necessário
        if usuarios[username]['tipo'] == 'colaborador':
            bordadores = carregar_bordadores()
            if old_name in bordadores:
                index = bordadores.index(old_name)
                bordadores[index] = data['nome']
                salvar_bordadores(bordadores)
            
            # Atualizar nos registros de produção
            dados = carregar_dados()
            for registro in dados:
                if registro.get('Bordador') == old_name:
                    registro['Bordador'] = data['nome']
            salvar_dados(dados)
    
    if 'senha' in data and data['senha']:
        from werkzeug.security import generate_password_hash
        usuarios[username]['senha'] = generate_password_hash(data['senha'])
    
    if 'tipo' in data:
        usuarios[username]['tipo'] = data['tipo']
    
    salvar_usuarios(usuarios)
    
    return jsonify({
        'success': True,
        'usuario': {
            'username': username,
            'nome': usuarios[username]['nome'],
            'tipo': usuarios[username]['tipo']
        }
    })

# Rota para excluir usuário
@app.route('/api/usuarios/<username>', methods=['DELETE'])
def delete_usuario(username):
    usuarios = carregar_usuarios()
    
    if username not in usuarios:
        return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404
    
    # Não permitir excluir o gerente principal
    if username == 'gerente':
        return jsonify({
            'success': False,
            'message': 'Não é possível excluir o usuário gerente principal'
        }), 400
    
    nome_usuario = usuarios[username]['nome']
    tipo_usuario = usuarios[username]['tipo']
    
    # Remover dos bordadores se for colaborador
    if tipo_usuario == 'colaborador':
        bordadores = carregar_bordadores()
        if nome_usuario in bordadores:
            bordadores.remove(nome_usuario)
            salvar_bordadores(bordadores)
    
    # Remover o usuário
    del usuarios[username]
    salvar_usuarios(usuarios)
    
    return jsonify({
        'success': True,
        'message': f'Usuário {username} excluído com sucesso'
    })

# Rota para alterar senha
@app.route('/api/usuarios/<username>/senha', methods=['PUT'])
def change_password(username):
    data = request.json
    usuarios = carregar_usuarios()
    
    if username not in usuarios:
        return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404
    
    nova_senha = data.get('senha')
    
    if not nova_senha or len(nova_senha) < 6:
        return jsonify({'success': False, 'message': 'Senha deve ter no mínimo 6 caracteres'}), 400
    
    from werkzeug.security import generate_password_hash
    usuarios[username]['senha'] = generate_password_hash(nova_senha)
    salvar_usuarios(usuarios)
    
    return jsonify({
        'success': True,
        'message': 'Senha alterada com sucesso'
    })


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)